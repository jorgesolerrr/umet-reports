import pandas as pd
from datetime import datetime

from src.utils.redis_client import RedisClient
from src.utils.logging.logger_factory import get_logger

logger = get_logger()


def get_grades_data(redis_client: RedisClient, patterns: list[str]) -> list[dict]:
    """
    Obtiene todos los datos de calificaciones almacenados en Redis.
    Como cada curso ya tiene sus filas del Excel completas, solo necesitamos
    combinar todas las filas de todos los cursos.
    
    Args:
        redis_client: Cliente de Redis
        patterns: Lista de patrones usados en la búsqueda
        
    Returns:
        list: Lista de filas para el Excel
    """
    all_excel_rows = []
    
    for pattern in patterns:
        logger.info(f"Obteniendo datos de Redis para patrón {pattern}")
        # Buscar todas las keys que empiecen con el patrón
        course_keys = redis_client.get_hset_keys_with_prefix(f"{pattern}:")
        logger.info(f"Encontradas {len(course_keys)} keys para patrón {pattern}")
        
        for course_key in course_keys:
            try:
                course_data = redis_client.get_hset(course_key)
                if course_data and "excel_rows" in course_data:
                    # Cada curso tiene una lista de filas ya completas
                    excel_rows = course_data["excel_rows"]
                    all_excel_rows.extend(excel_rows)
                    logger.info(f"Obtenidas {len(excel_rows)} filas de {course_key}")
            except Exception as e:
                logger.error(f"Error obteniendo datos de {course_key}: {e}")
                continue
    
    logger.info(f"Total de filas obtenidas de Redis: {len(all_excel_rows)}")
    return all_excel_rows


def make_grades_report(excel_rows: list[dict], filename: str = None) -> str:
    """
    Crea un archivo Excel con los datos del reporte de calificaciones.

    Args:
        excel_rows: Lista de diccionarios con las filas del reporte (ya completas)
        filename: Nombre del archivo. Si no se especifica, se genera automáticamente.

    Returns:
        str: Ruta del archivo Excel creado
    """
    if not excel_rows:
        logger.warning("No hay datos para generar el reporte")
        return None

    # Crear DataFrame
    df = pd.DataFrame(excel_rows)

    # Generar nombre del archivo si no se especifica
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_calificaciones_{timestamp}.xlsx"

    logger.info(f"Generando archivo Excel: {filename}")

    # Crear archivo Excel
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Calificaciones", index=False)
        
        # Formatear el Excel
        _format_excel_sheet(writer, df)

    logger.info(f"Archivo Excel generado exitosamente: {filename}")
    return filename


def _format_excel_sheet(writer, df):
    """
    Aplica formato al Excel para mejorar la presentación.

    Args:
        writer: ExcelWriter object
        df: DataFrame con los datos
    """
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    worksheet = writer.sheets["Calificaciones"]

    # Formatear encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Aplicar filtros
    if len(df) > 0:
        last_col = get_column_letter(len(df.columns))
        last_row = len(df) + 1
        worksheet.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Ajustar ancho de columnas
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
        worksheet.column_dimensions[column_letter].width = adjusted_width

