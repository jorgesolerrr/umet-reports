import json
import pandas as pd
from datetime import datetime

from src.utils.redis_client import RedisClient
from src.utils.logging.logger_factory import get_logger
from src.utils.category_finder import CategoryBuilder

logger = get_logger()


def get_courses_data(redis_client: RedisClient):
    result_contentTotals = []
    result_modulesIndexes = {}
    logger.info(
        f"Getting courses data from redis {redis_client.client.llen('total_report_json')}"
    )
    while redis_client.client.llen("total_report_json") > 0:
        slice_contentTotals_i = json.loads(
            redis_client.client.lpop("total_report_json")
        )
        slice_moduleIndexes_i = json.loads(redis_client.client.lpop("module_indexes"))
        if slice_contentTotals_i:
            result_contentTotals.append(slice_contentTotals_i)
        if slice_moduleIndexes_i:
            result_modulesIndexes |= slice_moduleIndexes_i
    return result_contentTotals, result_modulesIndexes


def parse_resources_totals(course_data: dict, params: dict, parcial_cut: int) -> dict:
    total_resources = (
        course_data["T_links_clases"]
        + course_data["T_links"]
        + course_data["T_archivos"]
        + course_data["T_label"]
        + course_data["T_pags"]
        + course_data["T_libros"]
    )
    current_cort_param = params[f"resources-cort-{parcial_cut}"]
    if total_resources >= current_cort_param:

        return {"TOTAL_RECURSOS": total_resources, "ESTADO_RECURSOS": "SUFICIENTE"}
    if total_resources >= current_cort_param - 3:
        return {"TOTAL_RECURSOS": total_resources, "ESTADO_RECURSOS": "INSUFICIENTE"}
    return {"TOTAL_RECURSOS": total_resources, "ESTADO_RECURSOS": "DEFICIENTE"}


def parse_activities_totals(course_data: dict, params: dict, parcial_cut: int) -> dict:
    total_activities = (
        course_data["T_tareas"]
        + course_data["T_quiz"]
        + course_data["T_glosario"]
        + course_data["T_taller"]
        + course_data["T_leccion"]
        + course_data["T_foro"]
        + course_data["T_wiki"]
        + course_data["T_chat"]
        + course_data["T_mapaMental"]
        
    )
    current_cort_param = params[f"activity-cort-{parcial_cut}"]
    if total_activities >= current_cort_param:
        return {
            "TOTAL_ACTIVIDADES": total_activities,
            "ESTADO_ACTIVIDADES": "SUFICIENTE",
            "ACTIVIDADES_ABIERTAS": course_data["T_actividades_abiertas"]
            
        }
    if total_activities >= current_cort_param - 3:
        return {
            "TOTAL_ACTIVIDADES": total_activities,
            "ESTADO_ACTIVIDADES": "INSUFICIENTE",
            "ACTIVIDADES_ABIERTAS": course_data["T_actividades_abiertas"]
        }
    return {
        "TOTAL_ACTIVIDADES": total_activities,
        "ESTADO_ACTIVIDADES": "DEFICIENTE",
        "ACTIVIDADES_ABIERTAS": course_data["T_actividades_abiertas"],
    }


def MakeQuantitativeReport(
    redis_client: RedisClient,
    category_builder: CategoryBuilder,
    params: dict,
    parcial_cut: int,
):
    rep_content, _ = get_courses_data(redis_client)
    logger.info("Iniciando reporte cuantitativo")
    courses_withoutTeacher = []
    courses_withProblems = []
    excel_rep = []
    for course in rep_content:
        if course["Total_Docentes"] == 0:
            courses_withoutTeacher.append(course["course_name"])
            profesor = None
        else:
            profesor = course["Docentes"].split(",")[0]
        try:
            row = {
                **category_builder.build_categories(course),
                "PROFESOR NOMBRE": profesor.split(":")[0] if profesor else "SIN PROFESOR",
                "PROFESOR CEDULA": profesor.split(":")[1] if profesor else "SIN PROFESOR",
                "NOMBRE": course["course_name"],
                "OFG": int(course["OFG"]),
                "CANT_ESTUDIANTES": course["Total_Estudiantes"],
                "secciones": course["Total_Secciones"],
                "links_clases": course["T_links_clases"],
                "links": course["T_links"],
                "archivos": course["T_archivos"],
                "labels": course["T_label"],
                "pags": course["T_pags"],
                "libros": course["T_libros"],
                **parse_resources_totals(course, params, parcial_cut),
                "tareas": course["T_tareas"],
                "quiz": course["T_quiz"],
                "glosario": course["T_glosario"],
                "taller": course["T_taller"],
                "leccion": course["T_leccion"],
                "foro": course["T_foro"],
                "wiki": course["T_wiki"],
                "chat": course["T_chat"],
                "mapaMental": course["T_mapaMental"],
                "CP1-ACT": course["CP1-ACT"],
                "CP1-ACTF": course["CP1-ACTF"],
                "CP2-ACT": course["CP2-ACT"],
                "CP2-ACTF": course["CP2-ACTF"],
                "CP3-ACT": course["CP3-ACT"],
                "CP3-ACTF": course["CP3-ACTF"],
                "CS-ACTF": course["CS-ACTF"],
                "CF-ACTF": course["CF-ACTF"],
                **parse_activities_totals(course, params, parcial_cut),
            }
            if row["TIPO APROBACION"] not in ["APROBACION REGULAR", "APROBACION TUTORIA"]:
                continue
            excel_rep.append(row)
        except Exception as e:
            logger.error(f"Error al procesar el curso {course['course_name']}: {e}")
            courses_withProblems.append(course["course_name"])
            continue
        del row
    logger.info("Reporte cuantitativo finalizado")
    return {
        "excel": excel_rep,
        "SIN PROFESORES": courses_withoutTeacher,
        "CON PROBLEMAS": courses_withProblems,
    }


def make_report(excel_rep: list, filename: str = None):
    """
    Crea un archivo Excel con múltiples hojas a partir de los datos del reporte cuantitativo.

    Args:
        excel_rep (list): Lista de diccionarios con los datos del reporte
        filename (str, optional): Nombre del archivo. Si no se especifica, se genera automáticamente.

    Returns:
        str: Ruta del archivo Excel creado
    """
    if not excel_rep:
        logger.warning("No hay datos para generar el reporte")
        return None

    # Crear DataFrame principal con todos los datos
    df_completo = pd.DataFrame(excel_rep)

    # Generar nombre del archivo si no se especifica
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_cuantitativo_{timestamp}.xlsx"

    logger.info(f"Generando archivo Excel: {filename}")

    # Crear archivo Excel con múltiples hojas
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:

        # Hoja 1: Datos Completos
        df_completo.to_excel(writer, sheet_name="Datos Completos", index=False)
        logger.info("Hoja 'Datos Completos' creada")

        # Hoja 2: Reporte de Recursos
        columnas_recursos = [
            "PROFESOR NOMBRE",
            "PROFESOR CEDULA",
            "NOMBRE",
            "OFG",
            "CANT_ESTUDIANTES",
            "links_clases",
            "links",
            "archivos",
            "labels",
            "pags",
            "libros",
            "TOTAL_RECURSOS",
            "ESTADO_RECURSOS",
        ]

        # Filtrar solo las columnas que existen en los datos
        columnas_recursos_existentes = [
            col for col in columnas_recursos if col in df_completo.columns
        ]
        df_recursos = df_completo[columnas_recursos_existentes].copy()

        # Ordenar por estado de recursos (primero DEFICIENTE, luego INSUFICIENTE, luego SUFICIENTE)
        order_recursos = ["DEFICIENTE", "INSUFICIENTE", "SUFICIENTE"]
        if "ESTADO_RECURSOS" in df_recursos.columns:
            df_recursos["_sort_order"] = df_recursos["ESTADO_RECURSOS"].map(
                {estado: i for i, estado in enumerate(order_recursos)}
            )
            df_recursos = df_recursos.sort_values("_sort_order").drop(
                "_sort_order", axis=1
            )

        df_recursos.to_excel(writer, sheet_name="Reporte Recursos", index=False)
        logger.info("Hoja 'Reporte Recursos' creada")

        # Hoja 3: Reporte de Actividades
        columnas_actividades = [
            "PROFESOR NOMBRE",
            "PROFESOR CEDULA",
            "NOMBRE",
            "OFG",
            "CANT_ESTUDIANTES",
            "tareas",
            "quiz",
            "glosario",
            "taller",
            "leccion",
            "foro",
            "wiki",
            "chat",
            "mapaMental",
            "TOTAL_ACTIVIDADES",
            "ESTADO_ACTIVIDADES",
            "ACTIVIDADES_ABIERTAS",
        ]

        # Filtrar solo las columnas que existen en los datos
        columnas_actividades_existentes = [
            col for col in columnas_actividades if col in df_completo.columns
        ]
        df_actividades = df_completo[columnas_actividades_existentes].copy()

        # Ordenar por estado de actividades (primero DEFICIENTE, luego INSUFICIENTE, luego SUFICIENTE)
        order_actividades = ["DEFICIENTE", "INSUFICIENTE", "SUFICIENTE"]
        if "ESTADO_ACTIVIDADES" in df_actividades.columns:
            df_actividades["_sort_order"] = df_actividades["ESTADO_ACTIVIDADES"].map(
                {estado: i for i, estado in enumerate(order_actividades)}
            )
            df_actividades = df_actividades.sort_values("_sort_order").drop(
                "_sort_order", axis=1
            )

        df_actividades.to_excel(writer, sheet_name="Reporte Actividades", index=False)
        logger.info("Hoja 'Reporte Actividades' creada")

        # Crear hojas de resumen por carrera y facultad
        create_summary_sheets(writer, df_completo)

        # Formatear las hojas para mejor presentación
        _format_excel_sheets(writer, df_completo, df_recursos, df_actividades)

    logger.info(f"Archivo Excel generado exitosamente: {filename}")
    return filename


def create_summary_sheets(writer, df_completo):
    """
    Crea hojas de resumen agrupadas por carrera y facultad.

    Args:
        writer: ExcelWriter object
        df_completo: DataFrame con todos los datos
    """
    # Verificar que existen las columnas necesarias
    required_cols = ["CARRERA", "FACULTAD", "ESTADO_RECURSOS", "ESTADO_ACTIVIDADES"]
    missing_cols = [col for col in required_cols if col not in df_completo.columns]

    if missing_cols:
        logger.warning(
            f"No se pueden crear hojas de resumen. Columnas faltantes: {missing_cols}"
        )
        return

    # Resumen por Carrera
    logger.info("Creando hoja de resumen por carrera")
    summary_carrera = _create_groupby_summary(df_completo, "CARRERA")
    summary_carrera.to_excel(writer, sheet_name="Resumen por Carrera", index=False)

    # Resumen por Facultad
    logger.info("Creando hoja de resumen por facultad")
    summary_facultad = _create_groupby_summary(df_completo, "FACULTAD")
    summary_facultad.to_excel(writer, sheet_name="Resumen por Facultad", index=False)


def _create_groupby_summary(df, group_column):
    """
    Crea un resumen agrupado por la columna especificada.

    Args:
        df: DataFrame con los datos
        group_column: Columna por la cual agrupar ('CARRERA' o 'FACULTAD')

    Returns:
        DataFrame con el resumen
    """
    # Crear tablas pivot para recursos y actividades
    recursos_pivot = pd.crosstab(df[group_column], df["ESTADO_RECURSOS"])
    actividades_pivot = pd.crosstab(df[group_column], df["ESTADO_ACTIVIDADES"])

    # Asegurar que todas las columnas de estado existen
    estados = ["DEFICIENTE", "INSUFICIENTE", "SUFICIENTE"]
    for estado in estados:
        if estado not in recursos_pivot.columns:
            recursos_pivot[estado] = 0
        if estado not in actividades_pivot.columns:
            actividades_pivot[estado] = 0

    # Reordenar columnas
    recursos_pivot = recursos_pivot[estados]
    actividades_pivot = actividades_pivot[estados]

    # Crear DataFrame de resumen
    summary_data = []

    for group_name in recursos_pivot.index:
        row = {
            group_column: group_name,
            "Total Cursos": recursos_pivot.loc[group_name].sum(),
            # Recursos
            "Recursos DEFICIENTE": recursos_pivot.loc[group_name, "DEFICIENTE"],
            "Recursos INSUFICIENTE": recursos_pivot.loc[group_name, "INSUFICIENTE"],
            "Recursos SUFICIENTE": recursos_pivot.loc[group_name, "SUFICIENTE"],
            # Actividades
            "Actividades DEFICIENTE": (
                actividades_pivot.loc[group_name, "DEFICIENTE"]
                if group_name in actividades_pivot.index
                else 0
            ),
            "Actividades INSUFICIENTE": (
                actividades_pivot.loc[group_name, "INSUFICIENTE"]
                if group_name in actividades_pivot.index
                else 0
            ),
            "Actividades SUFICIENTE": (
                actividades_pivot.loc[group_name, "SUFICIENTE"]
                if group_name in actividades_pivot.index
                else 0
            ),
        }

        # Calcular porcentajes
        total_cursos = row["Total Cursos"]
        if total_cursos > 0:
            row["% Recursos DEFICIENTE"] = round(
                (row["Recursos DEFICIENTE"] / total_cursos) * 100, 1
            )
            row["% Recursos INSUFICIENTE"] = round(
                (row["Recursos INSUFICIENTE"] / total_cursos) * 100, 1
            )
            row["% Recursos SUFICIENTE"] = round(
                (row["Recursos SUFICIENTE"] / total_cursos) * 100, 1
            )
            row["% Actividades DEFICIENTE"] = round(
                (row["Actividades DEFICIENTE"] / total_cursos) * 100, 1
            )
            row["% Actividades INSUFICIENTE"] = round(
                (row["Actividades INSUFICIENTE"] / total_cursos) * 100, 1
            )
            row["% Actividades SUFICIENTE"] = round(
                (row["Actividades SUFICIENTE"] / total_cursos) * 100, 1
            )
        else:
            for key in row.keys():
                if key.startswith("%"):
                    row[key] = 0.0

        summary_data.append(row)

    summary_df = pd.DataFrame(summary_data)

    # Ordenar por total de cursos (descendente)
    summary_df = summary_df.sort_values("Total Cursos", ascending=False)

    return summary_df


def _format_excel_sheets(writer, df_completo, df_recursos, df_actividades):
    """
    Aplica formato a las hojas del Excel para mejorar la presentación.

    Args:
        writer: ExcelWriter object
        df_completo: DataFrame con datos completos
        df_recursos: DataFrame con datos de recursos
        df_actividades: DataFrame con datos de actividades
    """
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    # Colores para los diferentes estados
    colors = {
        "SUFICIENTE": "C6EFCE",  # Verde claro
        "INSUFICIENTE": "FFEB9C",  # Amarillo claro
        "DEFICIENTE": "FFC7CE",  # Rojo claro
    }

    # Obtener todas las hojas del workbook
    all_sheets = list(writer.sheets.keys())

    # Formatear cada hoja
    for sheet_name in all_sheets:
        worksheet = writer.sheets[sheet_name]

        # Obtener el DataFrame correspondiente
        if sheet_name == "Datos Completos":
            df = df_completo
        elif sheet_name == "Reporte Recursos":
            df = df_recursos
        elif sheet_name == "Reporte Actividades":
            df = df_actividades
        else:
            # Para hojas de resumen, obtener el rango de datos
            max_row = worksheet.max_row
            max_col = worksheet.max_column

            # Aplicar filtros a los encabezados
            if max_row > 1 and max_col > 0:
                worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
                logger.info(f"Filtros aplicados a la hoja '{sheet_name}'")

            # Formatear encabezados para hojas de resumen
            _format_summary_headers(worksheet, max_col)
            continue

        # Formatear encabezados para hojas principales
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Aplicar filtros a los datos
        if len(df) > 0:
            last_col = get_column_letter(len(df.columns))
            last_row = len(df) + 1
            worksheet.auto_filter.ref = f"A1:{last_col}{last_row}"
            logger.info(f"Filtros aplicados a la hoja '{sheet_name}'")

        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Aplicar colores según el estado (solo para hojas de recursos y actividades)
        if sheet_name in ["Reporte Recursos", "Reporte Actividades"]:
            estado_col = (
                "ESTADO_RECURSOS"
                if "ESTADO_RECURSOS" in df.columns
                else "ESTADO_ACTIVIDADES"
            )

            if estado_col in df.columns:
                estado_col_index = df.columns.get_loc(estado_col) + 1

                for row_num in range(
                    2, len(df) + 2
                ):  # Empezar desde la fila 2 (después del encabezado)
                    estado_cell = worksheet.cell(row=row_num, column=estado_col_index)
                    estado_value = estado_cell.value

                    if estado_value in colors:
                        fill = PatternFill(
                            start_color=colors[estado_value],
                            end_color=colors[estado_value],
                            fill_type="solid",
                        )

                        # Aplicar color a toda la fila
                        for col_num in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row_num, column=col_num).fill = fill


def _format_summary_headers(worksheet, max_col):
    """
    Formatea los encabezados de las hojas de resumen.

    Args:
        worksheet: Hoja de trabajo de openpyxl
        max_col: Número máximo de columnas
    """
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Aplicar formato a la primera fila (encabezados)
    for col_num in range(1, max_col + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Ajustar ancho de columnas
    for col_num in range(1, max_col + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0

        for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
